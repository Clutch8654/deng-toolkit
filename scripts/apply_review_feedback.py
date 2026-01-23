#!/usr/bin/env python3
"""
Apply Review Feedback - Read corrections from Excel and update catalog.

Usage:
    python apply_review_feedback.py PATH_TO_EXCEL [--reviewer NAME]

Examples:
    python apply_review_feedback.py ~/Desktop/procedure_patterns_reviewed.xlsx
    python apply_review_feedback.py review.xlsx --reviewer "Jane Smith"
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

CATALOG_DIR = Path(os.environ.get("DENG_CATALOG_DIR", str(Path.home() / ".ds_catalog")))
REVIEWS_DIR = CATALOG_DIR / "reviews"
ONTOLOGY_PATH = CATALOG_DIR / "ontology.jsonld"
PROCEDURE_ANALYSIS_PATH = CATALOG_DIR / "procedure_analysis.jsonld"


def read_excel_feedback(excel_path: Path) -> dict:
    """Read the Excel file and extract all feedback."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(excel_path, data_only=True)

    feedback = {
        "metricCorrections": [],
        "relationshipFlags": [],
        "filterMeanings": [],
    }

    # Read Discovered Metrics sheet
    if "Discovered Metrics" in wb.sheetnames:
        ws = wb["Discovered Metrics"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 9:
                continue

            correct_name = row[7]  # Column H: Correct Name
            notes = row[8]  # Column I: Notes

            # Skip if no feedback provided
            if not correct_name and not notes:
                continue

            feedback["metricCorrections"].append(
                {
                    "procedure": row[3],  # Column D: Procedure
                    "foundName": row[4],  # Column E: Metric Name
                    "correctedName": correct_name or None,
                    "notes": notes or None,
                    "formula": row[5],  # Column F: Formula (for reference)
                }
            )

    # Read Table Relationships sheet
    if "Table Relationships" in wb.sheetnames:
        ws = wb["Table Relationships"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 8:
                continue

            is_correct = row[6]  # Column G: Is Correct?
            notes = row[7]  # Column H: Notes

            # Skip if no feedback provided
            if not is_correct and not notes:
                continue

            # Convert is_correct to boolean
            is_correct_bool = None
            if is_correct:
                is_correct_lower = str(is_correct).lower().strip()
                if is_correct_lower == "yes":
                    is_correct_bool = True
                elif is_correct_lower == "no":
                    is_correct_bool = False
                # "unsure" stays as None

            feedback["relationshipFlags"].append(
                {
                    "procedure": row[1],  # Column B: Procedure
                    "leftTable": row[2],  # Column C: Left Table
                    "rightTable": row[3],  # Column D: Right Table
                    "joinType": row[4],  # Column E: Join Type
                    "isCorrect": is_correct_bool,
                    "notes": notes or None,
                }
            )

    # Read Common Filters sheet
    if "Common Filters" in wb.sheetnames:
        ws = wb["Common Filters"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 6:
                continue

            business_meaning = row[5]  # Column F: Business Meaning

            # Skip if no feedback provided
            if not business_meaning:
                continue

            feedback["filterMeanings"].append(
                {
                    "column": row[1],  # Column B: Column
                    "operator": row[2],  # Column C: Operator
                    "values": row[3],  # Column D: Typical Values
                    "businessMeaning": business_meaning,
                }
            )

    return feedback


def save_feedback_json(feedback: dict, excel_path: Path, reviewer: str | None) -> Path:
    """Save the feedback as JSON."""
    date_str = datetime.now().strftime("%Y-%m-%d")
    output_path = REVIEWS_DIR / f"procedure_feedback_{date_str}.json"

    full_feedback = {
        "reviewedBy": reviewer,
        "reviewedAt": datetime.now().isoformat(),
        "sourceFile": str(excel_path.name),
        **feedback,
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w") as f:
        json.dump(full_feedback, f, indent=2)

    return output_path


def apply_to_ontology(feedback: dict) -> int:
    """Apply feedback to the ontology.jsonld file."""
    if not ONTOLOGY_PATH.exists():
        print("WARNING: ontology.jsonld not found, skipping ontology update")
        return 0

    with open(ONTOLOGY_PATH) as f:
        ontology = json.load(f)

    changes = 0

    # Apply metric corrections - add businessTerms to columns
    metric_map = {}
    for mc in feedback["metricCorrections"]:
        if mc["correctedName"]:
            # Map formula columns to business names
            for col in _extract_columns_from_formula(mc.get("formula", "")):
                metric_map[col] = mc["correctedName"]

    # Apply filter meanings - add businessMeaning to columns
    filter_map = {}
    for fm in feedback["filterMeanings"]:
        col = fm["column"]
        if "." in col:
            col = col.split(".")[-1]  # Get just column name
        filter_map[col] = fm["businessMeaning"]

    # Update ontology entities
    for entity in ontology.get("entities", []):
        for col in entity.get("hasColumn", []):
            col_name = col.get("rdfs:label", "")

            # Apply metric term
            if col_name in metric_map:
                if "businessTerms" not in col:
                    col["businessTerms"] = []
                if metric_map[col_name] not in col["businessTerms"]:
                    col["businessTerms"].append(metric_map[col_name])
                    changes += 1

            # Apply filter meaning
            if col_name in filter_map:
                col["businessMeaning"] = filter_map[col_name]
                changes += 1

    # Save updated ontology
    if changes > 0:
        with open(ONTOLOGY_PATH, "w") as f:
            json.dump(ontology, f, indent=2)
        print(f"Updated ontology with {changes} changes")

    return changes


def apply_to_procedure_analysis(feedback: dict) -> int:
    """Apply feedback to the procedure_analysis.jsonld file."""
    if not PROCEDURE_ANALYSIS_PATH.exists():
        print("WARNING: procedure_analysis.jsonld not found, skipping update")
        return 0

    with open(PROCEDURE_ANALYSIS_PATH) as f:
        analysis = json.load(f)

    changes = 0

    # Add a reviewFeedback section
    if "reviewFeedback" not in analysis:
        analysis["reviewFeedback"] = {
            "metricCorrections": [],
            "relationshipFlags": [],
            "filterMeanings": [],
        }

    # Merge feedback
    for mc in feedback["metricCorrections"]:
        analysis["reviewFeedback"]["metricCorrections"].append(mc)
        changes += 1

    for rf in feedback["relationshipFlags"]:
        analysis["reviewFeedback"]["relationshipFlags"].append(rf)
        changes += 1

    for fm in feedback["filterMeanings"]:
        analysis["reviewFeedback"]["filterMeanings"].append(fm)
        changes += 1

    # Save
    if changes > 0:
        with open(PROCEDURE_ANALYSIS_PATH, "w") as f:
            json.dump(analysis, f, indent=2, default=str)
        print(f"Updated procedure_analysis.jsonld with {changes} feedback items")

    return changes


def _extract_columns_from_formula(formula: str) -> list[str]:
    """Extract column names from a SQL formula (simple extraction)."""
    if not formula:
        return []

    # Simple heuristic: find words that look like column names
    import re

    # Match table.column or just column patterns
    matches = re.findall(r"(?:\w+\.)?(\w+)", formula)
    # Filter out SQL keywords
    keywords = {
        "case",
        "when",
        "then",
        "else",
        "end",
        "and",
        "or",
        "not",
        "sum",
        "count",
        "avg",
        "min",
        "max",
        "nullif",
        "cast",
        "as",
    }
    return [m for m in matches if m.lower() not in keywords and not m.isdigit()]


def print_feedback_summary(feedback: dict) -> None:
    """Print a summary of the feedback found."""
    print("\n=== Feedback Summary ===")

    print(f"\nMetric Corrections: {len(feedback['metricCorrections'])}")
    for mc in feedback["metricCorrections"][:5]:
        print(f"  - {mc['foundName']} -> {mc['correctedName'] or '(note only)'}")
        if mc.get("notes"):
            print(f"    Note: {mc['notes'][:60]}...")
    if len(feedback["metricCorrections"]) > 5:
        print(f"  ... and {len(feedback['metricCorrections']) - 5} more")

    print(f"\nRelationship Flags: {len(feedback['relationshipFlags'])}")
    for rf in feedback["relationshipFlags"][:5]:
        status = (
            "Correct"
            if rf["isCorrect"]
            else ("Incorrect" if rf["isCorrect"] is False else "Unsure")
        )
        print(f"  - {rf['leftTable']} -> {rf['rightTable']}: {status}")
        if rf.get("notes"):
            print(f"    Note: {rf['notes'][:60]}...")
    if len(feedback["relationshipFlags"]) > 5:
        print(f"  ... and {len(feedback['relationshipFlags']) - 5} more")

    print(f"\nFilter Meanings: {len(feedback['filterMeanings'])}")
    for fm in feedback["filterMeanings"][:5]:
        print(f"  - {fm['column']} {fm['operator']}: {fm['businessMeaning'][:50]}")
    if len(feedback["filterMeanings"]) > 5:
        print(f"  ... and {len(feedback['filterMeanings']) - 5} more")


def main():
    parser = argparse.ArgumentParser(
        description="Apply review feedback from Excel to catalog"
    )
    parser.add_argument(
        "excel_path",
        type=Path,
        help="Path to the reviewed Excel file",
    )
    parser.add_argument(
        "--reviewer",
        help="Name of the reviewer (optional)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Show what would be applied without making changes",
    )

    args = parser.parse_args()

    if not args.excel_path.exists():
        print(f"ERROR: File not found: {args.excel_path}")
        sys.exit(1)

    # Read feedback from Excel
    print(f"Reading feedback from: {args.excel_path}")
    feedback = read_excel_feedback(args.excel_path)

    # Show summary
    print_feedback_summary(feedback)

    total_feedback = (
        len(feedback["metricCorrections"])
        + len(feedback["relationshipFlags"])
        + len(feedback["filterMeanings"])
    )

    if total_feedback == 0:
        print("\nNo feedback found in the Excel file.")
        print("Make sure you filled in the yellow columns.")
        sys.exit(0)

    if args.dry_run:
        print(
            "\n[DRY RUN] Would apply the above feedback. Run without --dry-run to apply."
        )
        sys.exit(0)

    # Save feedback JSON
    json_path = save_feedback_json(feedback, args.excel_path, args.reviewer)
    print(f"\nSaved feedback JSON to: {json_path}")

    # Apply to ontology
    ontology_changes = apply_to_ontology(feedback)

    # Apply to procedure analysis
    analysis_changes = apply_to_procedure_analysis(feedback)

    print(f"\n=== Feedback Applied ===")
    print(f"  Feedback items: {total_feedback}")
    print(f"  Ontology updates: {ontology_changes}")
    print(f"  Analysis updates: {analysis_changes}")


if __name__ == "__main__":
    main()
