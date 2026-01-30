#!/usr/bin/env python3
"""
Generate Excel Review File - Create spreadsheet for business user review of parsed SQL patterns.

Usage:
    python generate_review_excel.py [--top N] [--output PATH]

Examples:
    python generate_review_excel.py --top 50
    python generate_review_excel.py --output ~/Desktop/review.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# Excel library - required for output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.worksheet.datavalidation import DataValidation

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from config import get_catalog_dir

CATALOG_DIR = get_catalog_dir()
REVIEWS_DIR = CATALOG_DIR / "reviews"
PROCEDURES_PATH = CATALOG_DIR / "procedures.parquet"
EXECUTION_STATS_PATH = CATALOG_DIR / "execution_stats.parquet"
PROCEDURE_ANALYSIS_PATH = CATALOG_DIR / "procedure_analysis.jsonld"


def load_parsed_analysis() -> dict:
    """Load the full procedure analysis including pre-parsed patterns."""
    if not PROCEDURE_ANALYSIS_PATH.exists():
        return {"procedures": [], "globalPatterns": {}}

    with open(PROCEDURE_ANALYSIS_PATH) as f:
        return json.load(f)


def get_procedures_with_patterns(analysis: dict, top_n: int) -> list[dict]:
    """Get top N procedures with their pre-parsed patterns."""
    procedures = analysis.get("procedures", [])
    # Already sorted by importance in analyze_procedures.py
    return procedures[:top_n]


def build_metrics_data(procedures: list[dict], global_patterns: dict) -> list[dict]:
    """Extract metrics from pre-parsed procedures."""
    metrics = []

    for proc in procedures:
        patterns = proc.get("parsedPatterns", {})
        proc_metrics = patterns.get("metrics", [])

        for metric in proc_metrics:
            columns_used = metric.get("columnsUsed", [])
            metrics.append(
                {
                    "importanceScore": proc.get("importanceScore", 0),
                    "executionCount": proc.get("executionCount", 0),
                    "lastExecuted": proc.get("lastExecuted"),
                    "procedure": f"{proc.get('database', '')}.{proc.get('objectName', '')}",
                    "metricName": metric.get("name") or "(unnamed)",
                    "formula": (metric.get("formula") or "")[:200],
                    "columnsUsed": ", ".join(columns_used[:5]) if columns_used else "",
                    "isRatio": "Yes" if metric.get("isRatio") else "No",
                    "plainEnglish": _interpret_metric_dict(metric),
                    "correctName": "",
                    "notes": "",
                }
            )

    # Also include metrics from globalPatterns if available
    for metric in global_patterns.get("discoveredMetrics", []):
        if not any(m["formula"] == metric.get("formula") for m in metrics):
            columns_used = metric.get("columnsUsed", [])
            metrics.append(
                {
                    "importanceScore": 0,
                    "executionCount": 0,
                    "lastExecuted": None,
                    "procedure": metric.get("proc", ""),
                    "metricName": metric.get("name") or "(unnamed)",
                    "formula": (metric.get("formula") or "")[:200],
                    "columnsUsed": ", ".join(columns_used[:5]) if columns_used else "",
                    "isRatio": "Yes" if metric.get("isRatio") else "No",
                    "plainEnglish": _interpret_metric_dict(metric),
                    "correctName": "",
                    "notes": "",
                }
            )

    metrics.sort(key=lambda x: x["importanceScore"], reverse=True)
    return metrics


def _interpret_metric_dict(metric: dict) -> str:
    """Generate business interpretation of a metric."""
    name = (metric.get("name") or "").lower()
    formula = (metric.get("formula") or "").lower()
    cols = metric.get("columnsUsed", [])
    cols_lower = [c.lower() for c in cols]

    # Revenue/financial metrics
    if any(
        x in name or x in formula
        for x in ["revenue", "charge", "amount", "price", "cost"]
    ):
        if "net" in name or "net" in formula:
            return "Net revenue calculation (after discounts/adjustments)"
        if "discount" in name or "discount" in formula:
            return "Discount amount or rate calculation"
        if metric.get("isRatio"):
            return "Financial ratio (e.g., margin, discount rate)"
        return "Revenue or monetary calculation"

    # Count/volume metrics
    if "count" in formula:
        if any(x in formula for x in ["cncl", "cancel"]):
            return "Count of cancelled items - potential churn metric"
        if any(x in formula for x in ["fulf", "fulfill", "complete"]):
            return "Count of fulfilled/completed items"
        if "order" in formula or "order" in name:
            return "Order volume count"
        if "invoice" in formula:
            return "Invoice count"
        return "Volume/count metric"

    # Rate/percentage metrics
    if metric.get("isRatio"):
        if any(x in cols_lower for x in ["statuscode", "status"]):
            return "Status-based rate (e.g., churn rate, fulfillment rate)"
        if any(x in name for x in ["rate", "percent", "pct", "ratio"]):
            return "Percentage or rate calculation"
        return "Ratio calculation - likely a KPI"

    # Status-based metrics
    if any(x in cols_lower for x in ["statuscode", "status"]):
        return "Status-based metric (tracks order/item states)"

    # Date-based metrics
    if any(x in cols_lower for x in ["date", "time", "created", "modified"]):
        return "Time-based calculation (duration, aging, etc.)"

    # Default interpretation based on name
    if name and name != "(unnamed)":
        readable_name = name.replace("_", " ").title()
        return f"Business metric: {readable_name}"

    return "Calculated business metric"


def _interpret_join(left_table: str, right_table: str, join_cols: list) -> str:
    """Generate business interpretation of a table join."""
    left = left_table.lower().split(".")[-1] if left_table else ""
    right = right_table.lower().split(".")[-1] if right_table else ""

    # Common join patterns
    patterns = {
        ("order", "orderitem"): "Order to line items - order details breakdown",
        ("orderitem", "order"): "Line items to parent order",
        ("order", "company"): "Orders to customer company",
        ("order", "account"): "Orders to customer account",
        ("order", "property"): "Orders to property location",
        ("orderitem", "product"): "Line items to product catalog",
        ("invoice", "invoiceitem"): "Invoice to line items - billing details",
        ("invoiceitem", "invoice"): "Invoice items to parent invoice",
        ("quote", "quoteitem"): "Quote to line items - proposal details",
        ("quoteitem", "quote"): "Quote items to parent quote",
        ("company", "property"): "Company to managed properties",
        ("property", "company"): "Properties to owning company",
        ("product", "family"): "Product to product family/category",
        ("product", "charge"): "Product to pricing/charges",
    }

    # Check for known patterns
    for (l, r), desc in patterns.items():
        if l in left and r in right:
            return desc
        if l in right and r in left:
            return desc

    # Infer from table names
    if "item" in right:
        return f"Parent to child items relationship"
    if "item" in left:
        return f"Child items to parent relationship"
    if any(x in right for x in ["company", "customer", "account"]):
        return f"Links to customer/account information"
    if any(x in right for x in ["product", "charge", "price"]):
        return f"Links to product/pricing information"
    if any(x in right for x in ["property", "location", "site"]):
        return f"Links to property/location information"

    return "Table relationship"


def _interpret_aggregation(func: str, column: str, alias: str) -> str:
    """Generate business interpretation of an aggregation."""
    func = (func or "").upper()
    col = (column or "").lower()
    alias_lower = (alias or "").lower()

    # COUNT interpretations
    if func == "COUNT":
        if col == "*":
            if "order" in alias_lower:
                return "Total order count"
            if "invoice" in alias_lower:
                return "Total invoice count"
            if "item" in alias_lower:
                return "Total line item count"
            return "Record count - volume metric"
        if "status" in col:
            return "Count by status - for status distribution analysis"
        if "id" in col:
            return "Distinct entity count"
        return f"Count of {col.replace('_', ' ')}"

    # SUM interpretations
    if func == "SUM":
        if any(x in col for x in ["amount", "charge", "price", "cost", "revenue"]):
            if "net" in col:
                return "Total net revenue (after adjustments)"
            if "discount" in col:
                return "Total discounts applied"
            return "Total monetary value - financial metric"
        if "quantity" in col or "qty" in col or "units" in col:
            return "Total quantity/units"
        return f"Sum of {col.replace('_', ' ')}"

    # AVG interpretations
    if func == "AVG":
        if any(x in col for x in ["amount", "charge", "price"]):
            return "Average transaction value"
        if "discount" in col:
            return "Average discount rate"
        if "duration" in col or "days" in col:
            return "Average time duration"
        return f"Average {col.replace('_', ' ')}"

    # MIN/MAX interpretations
    if func in ("MIN", "MAX"):
        if "date" in col or "time" in col:
            return f"{'Earliest' if func == 'MIN' else 'Latest'} date/time"
        if any(x in col for x in ["amount", "charge", "price"]):
            return f"{'Minimum' if func == 'MIN' else 'Maximum'} value"
        return f"{func} of {col.replace('_', ' ')}"

    return f"{func}({column})"


def _interpret_filter(column: str, operator: str, values: str) -> str:
    """Generate business interpretation of a filter condition."""
    col = (column or "").lower().split(".")[-1]  # Remove table prefix
    op = (operator or "").upper()
    vals = (values or "").upper()

    # Status code filters
    if "status" in col:
        if "CNCL" in vals:
            return "Filtering to CANCELLED items - churn/cancellation analysis"
        if "FULF" in vals:
            return "Filtering to FULFILLED items - completed orders"
        if "PEND" in vals:
            return "Filtering to PENDING items - in-progress orders"
        if "EXPD" in vals:
            return "Filtering to EXPIRED items"
        if op == "IN":
            return "Filtering by specific status codes"
        return "Status-based filter"

    # Date filters
    if any(x in col for x in ["date", "time", "created", "modified"]):
        if op in (">", ">=", "BETWEEN"):
            return "Date range filter - time-bounded query"
        return "Date-based filter"

    # ID filters (likely parameter-driven)
    if col.endswith("idseq") or col.endswith("id"):
        if "@" in vals.lower():
            return "Parameter-driven lookup by ID"
        return "Filtering by specific entity ID"

    # Flag/boolean filters
    if any(x in col for x in ["flag", "is_", "has_"]):
        return "Boolean/flag filter"

    # Type filters
    if "type" in col or "code" in col:
        return "Filtering by type/category code"

    # NULL checks
    if op == "IS NULL":
        return f"Finding records where {col.replace('_', ' ')} is missing"
    if op == "IS NOT NULL":
        return f"Finding records where {col.replace('_', ' ')} exists"

    return f"Filter on {col.replace('_', ' ')}"


def build_joins_data(procedures: list[dict]) -> list[dict]:
    """Extract join patterns from pre-parsed procedures."""
    joins = []

    for proc in procedures:
        patterns = proc.get("parsedPatterns", {})
        proc_joins = patterns.get("joins", [])

        for join in proc_joins:
            on_columns = join.get("onColumns", [])
            left_table = join.get("leftTable", "UNKNOWN")
            right_table = join.get("rightTable", "UNKNOWN")

            joins.append(
                {
                    "importanceScore": proc.get("importanceScore", 0),
                    "procedure": f"{proc.get('database', '')}.{proc.get('objectName', '')}",
                    "leftTable": left_table,
                    "rightTable": right_table,
                    "joinType": join.get("joinType", "INNER"),
                    "joinColumns": " = ".join([f"{c[0]} = {c[1]}" for c in on_columns])
                    if on_columns
                    else "",
                    "businessMeaning": _interpret_join(
                        left_table, right_table, on_columns
                    ),
                    "isCorrect": "",
                    "notes": "",
                }
            )

    joins.sort(key=lambda x: x["importanceScore"], reverse=True)
    return joins


def build_aggregations_data(procedures: list[dict]) -> list[dict]:
    """Extract aggregation patterns from pre-parsed procedures."""
    aggregations = []

    for proc in procedures:
        patterns = proc.get("parsedPatterns", {})
        proc_aggs = patterns.get("aggregations", [])

        for agg in proc_aggs:
            func = agg.get("function", "")
            column = agg.get("column", "")
            alias = agg.get("alias") or ""

            aggregations.append(
                {
                    "importanceScore": proc.get("importanceScore", 0),
                    "procedure": f"{proc.get('database', '')}.{proc.get('objectName', '')}",
                    "function": func,
                    "column": column,
                    "alias": alias,
                    "context": agg.get("context") or "",
                    "businessMeaning": _interpret_aggregation(func, column, alias),
                }
            )

    aggregations.sort(key=lambda x: x["importanceScore"], reverse=True)
    return aggregations


def build_filters_data(procedures: list[dict]) -> list[dict]:
    """Extract and aggregate filter patterns from pre-parsed procedures."""
    filter_counts = defaultdict(lambda: {"count": 0, "procedures": [], "importance": 0})

    for proc in procedures:
        patterns = proc.get("parsedPatterns", {})
        proc_filters = patterns.get("filters", [])
        proc_name = proc.get("objectName", "")
        importance = proc.get("importanceScore", 0)

        for flt in proc_filters:
            key = (
                flt.get("column", ""),
                flt.get("operator", ""),
                flt.get("valuePattern", ""),
            )
            filter_counts[key]["count"] += 1
            filter_counts[key]["importance"] = max(
                filter_counts[key]["importance"], importance
            )
            if len(filter_counts[key]["procedures"]) < 3:
                filter_counts[key]["procedures"].append(proc_name)

    filters = []
    for (column, operator, value_pattern), data in filter_counts.items():
        filters.append(
            {
                "frequency": data["count"],
                "importance": data["importance"],
                "column": column,
                "operator": operator,
                "typicalValues": (value_pattern or "")[:100],
                "usedIn": ", ".join(data["procedures"]),
                "businessMeaning": _interpret_filter(column, operator, value_pattern),
            }
        )

    filters.sort(key=lambda x: (-x["frequency"], -x["importance"]))
    return filters


def create_excel(
    metrics: list[dict],
    joins: list[dict],
    aggregations: list[dict],
    filters: list[dict],
    output_path: Path,
) -> None:
    """Create the Excel workbook with all sheets."""
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font_white = Font(bold=True, color="FFFFFF")
    editable_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Sheet 1: Instructions
    ws_inst = wb.active
    ws_inst.title = "Instructions"

    instructions = [
        ["Procedure Analysis Review"],
        [""],
        ["Purpose:"],
        [
            "This workbook contains patterns discovered by analyzing SQL stored procedures."
        ],
        ["Your feedback helps us understand the business meaning of these patterns."],
        [""],
        ["Sheets:"],
        [
            "- Discovered Metrics: Calculations and formulas found in SQL (ratios, percentages)"
        ],
        ["- Table Relationships: JOIN patterns between tables"],
        ["- Aggregations: SUM, COUNT, AVG operations - what's being measured"],
        ["- Common Filters: WHERE clause patterns (business rules)"],
        [""],
        ["How to Review:"],
        ["1. Yellow columns are for your input"],
        ["2. Items are sorted by importance (most-used procedures first)"],
        ["3. Focus on the top items - those have the most business impact"],
        ["4. Add corrections or clarifications in the yellow columns"],
        [""],
        ["When Done:"],
        [
            "1. Add your name to the filename (e.g., procedure_patterns_2026-01-22_JaneSmith.xlsx)"
        ],
        ["2. Email the file to: jensen.carlsen@realpage.com"],
        [""],
        [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
    ]

    for row in instructions:
        ws_inst.append(row)

    ws_inst["A1"].font = Font(bold=True, size=14)
    ws_inst.column_dimensions["A"].width = 80

    # Sheet 2: Discovered Metrics
    ws_metrics = wb.create_sheet("Discovered Metrics")

    metrics_headers = [
        "Importance Score",
        "Execution Count",
        "Last Executed",
        "Procedure",
        "Metric Name",
        "Formula",
        "Plain English",
        "Correct Name",
        "Notes",
    ]
    ws_metrics.append(metrics_headers)

    for col_idx, header in enumerate(metrics_headers, 1):
        cell = ws_metrics.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        # Mark editable columns
        if header in ["Correct Name", "Notes"]:
            cell.fill = editable_fill
            cell.font = header_font

    for metric in metrics:
        ws_metrics.append(
            [
                metric["importanceScore"],
                metric["executionCount"],
                metric["lastExecuted"] or "",
                metric["procedure"],
                metric["metricName"],
                metric["formula"],
                metric["plainEnglish"],
                metric["correctName"],
                metric["notes"],
            ]
        )

    # Apply editable fill to data cells in editable columns
    for row_idx in range(2, len(metrics) + 2):
        ws_metrics.cell(row=row_idx, column=8).fill = editable_fill  # Correct Name
        ws_metrics.cell(row=row_idx, column=9).fill = editable_fill  # Notes

    # Column widths
    ws_metrics.column_dimensions["A"].width = 15
    ws_metrics.column_dimensions["B"].width = 15
    ws_metrics.column_dimensions["C"].width = 20
    ws_metrics.column_dimensions["D"].width = 35
    ws_metrics.column_dimensions["E"].width = 25
    ws_metrics.column_dimensions["F"].width = 50
    ws_metrics.column_dimensions["G"].width = 40
    ws_metrics.column_dimensions["H"].width = 25
    ws_metrics.column_dimensions["I"].width = 30

    ws_metrics.freeze_panes = "A2"

    # Sheet 3: Table Relationships
    ws_joins = wb.create_sheet("Table Relationships")

    joins_headers = [
        "Importance Score",
        "Procedure",
        "Left Table",
        "Right Table",
        "Join Type",
        "Join Columns",
        "Business Interpretation",
        "Is Correct?",
        "Notes",
    ]
    ws_joins.append(joins_headers)

    for col_idx, header in enumerate(joins_headers, 1):
        cell = ws_joins.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        if header in ["Is Correct?", "Notes"]:
            cell.fill = editable_fill
            cell.font = header_font

    for join in joins:
        ws_joins.append(
            [
                join["importanceScore"],
                join["procedure"],
                join["leftTable"],
                join["rightTable"],
                join["joinType"],
                join["joinColumns"],
                join.get("businessMeaning", ""),
                join["isCorrect"],
                join["notes"],
            ]
        )

    # Add dropdown for Is Correct?
    dv = DataValidation(type="list", formula1='"Yes,No,Unsure"', allow_blank=True)
    dv.error = "Please select from dropdown"
    dv.errorTitle = "Invalid input"
    ws_joins.add_data_validation(dv)

    for row_idx in range(2, len(joins) + 2):
        ws_joins.cell(row=row_idx, column=8).fill = editable_fill  # Is Correct?
        ws_joins.cell(row=row_idx, column=9).fill = editable_fill  # Notes
        dv.add(ws_joins.cell(row=row_idx, column=8))

    ws_joins.column_dimensions["A"].width = 15
    ws_joins.column_dimensions["B"].width = 35
    ws_joins.column_dimensions["C"].width = 25
    ws_joins.column_dimensions["D"].width = 25
    ws_joins.column_dimensions["E"].width = 12
    ws_joins.column_dimensions["F"].width = 35
    ws_joins.column_dimensions["G"].width = 40  # Business Interpretation
    ws_joins.column_dimensions["H"].width = 12
    ws_joins.column_dimensions["I"].width = 30

    ws_joins.freeze_panes = "A2"

    # Sheet 4: Aggregations
    ws_aggs = wb.create_sheet("Aggregations")

    aggs_headers = [
        "Importance Score",
        "Procedure",
        "Function",
        "Column",
        "Alias",
        "Interpretation",
        "Your Correction",
    ]
    ws_aggs.append(aggs_headers)

    for col_idx, header in enumerate(aggs_headers, 1):
        cell = ws_aggs.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        if header == "Your Correction":
            cell.fill = editable_fill
            cell.font = header_font

    for agg in aggregations:
        ws_aggs.append(
            [
                agg["importanceScore"],
                agg["procedure"],
                agg["function"],
                agg["column"],
                agg["alias"],
                agg["businessMeaning"],  # Auto-generated interpretation
                "",  # Your Correction - for reviewer
            ]
        )

    for row_idx in range(2, len(aggregations) + 2):
        ws_aggs.cell(row=row_idx, column=7).fill = editable_fill

    ws_aggs.column_dimensions["A"].width = 15
    ws_aggs.column_dimensions["B"].width = 35
    ws_aggs.column_dimensions["C"].width = 10
    ws_aggs.column_dimensions["D"].width = 25
    ws_aggs.column_dimensions["E"].width = 20
    ws_aggs.column_dimensions["F"].width = 45  # Interpretation
    ws_aggs.column_dimensions["G"].width = 35  # Your Correction

    ws_aggs.freeze_panes = "A2"

    # Sheet 5: Common Filters
    ws_filters = wb.create_sheet("Common Filters")

    filters_headers = [
        "Frequency",
        "Column",
        "Operator",
        "Typical Values",
        "Used In",
        "Interpretation",
        "Your Correction",
    ]
    ws_filters.append(filters_headers)

    for col_idx, header in enumerate(filters_headers, 1):
        cell = ws_filters.cell(row=1, column=col_idx)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        if header == "Your Correction":
            cell.fill = editable_fill
            cell.font = header_font

    for flt in filters:
        ws_filters.append(
            [
                flt["frequency"],
                flt["column"],
                flt["operator"],
                flt["typicalValues"],
                flt["usedIn"],
                flt["businessMeaning"],  # Auto-generated interpretation
                "",  # Your Correction - for reviewer
            ]
        )

    for row_idx in range(2, len(filters) + 2):
        ws_filters.cell(row=row_idx, column=7).fill = editable_fill

    ws_filters.column_dimensions["A"].width = 12
    ws_filters.column_dimensions["B"].width = 30
    ws_filters.column_dimensions["C"].width = 12
    ws_filters.column_dimensions["D"].width = 35
    ws_filters.column_dimensions["E"].width = 35
    ws_filters.column_dimensions["F"].width = 45  # Interpretation
    ws_filters.column_dimensions["G"].width = 35  # Your Correction

    ws_filters.freeze_panes = "A2"

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\nExcel file saved to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel review file for procedure analysis"
    )
    parser.add_argument(
        "--top",
        type=int,
        default=100,
        help="Number of top procedures to include (default: 100)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output path (default: ~/.ds_catalog/reviews/procedure_patterns_YYYY-MM-DD.xlsx)",
    )

    args = parser.parse_args()

    # Check required files exist
    if not PROCEDURE_ANALYSIS_PATH.exists():
        print(f"ERROR: Procedure analysis not found: {PROCEDURE_ANALYSIS_PATH}")
        print("Run analyze_procedures.py --deep-analysis first.")
        sys.exit(1)

    # Load pre-parsed analysis
    print("Loading procedure analysis...")
    analysis = load_parsed_analysis()
    global_patterns = analysis.get("globalPatterns", {})

    procedures = get_procedures_with_patterns(analysis, args.top)
    print(f"  Found {len(procedures)} procedures with parsed patterns")

    # Check if we have parsed data
    has_patterns = any(proc.get("parsedPatterns") for proc in procedures)
    if not has_patterns:
        print(
            "WARNING: No parsed patterns found. Run analyze_procedures.py --deep-analysis first."
        )

    # Build data for sheets
    print("\nBuilding review data...")
    metrics = build_metrics_data(procedures, global_patterns)
    print(f"  Metrics: {len(metrics)}")

    joins = build_joins_data(procedures)
    print(f"  Join patterns: {len(joins)}")

    aggregations = build_aggregations_data(procedures)
    print(f"  Aggregations: {len(aggregations)}")

    filters = build_filters_data(procedures)
    print(f"  Filter patterns: {len(filters)}")

    # Generate Excel
    if args.output:
        output_path = args.output
    else:
        date_str = datetime.now().strftime("%Y-%m-%d")
        output_path = REVIEWS_DIR / f"procedure_patterns_{date_str}.xlsx"

    print("\nGenerating Excel...")
    create_excel(metrics, joins, aggregations, filters, output_path)

    print(f"\n=== Review File Generated ===")
    print(f"  Metrics: {len(metrics)}")
    print(f"  Joins: {len(joins)}")
    print(f"  Aggregations: {len(aggregations)}")
    print(f"  Filters: {len(filters)}")
    print(f"\nShare this file with reviewers:")
    print(f"  {output_path}")


if __name__ == "__main__":
    main()
